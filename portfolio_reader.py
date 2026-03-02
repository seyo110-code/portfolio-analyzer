"""
포트폴리오 엑셀 파일 읽기 및 구조화
- portfolio.xlsx 를 읽어서 분석 가능한 형태로 변환
- 평가손익/수익률/비중 자동 계산 후 엑셀에 반영
"""
import pandas as pd
import openpyxl
from pathlib import Path

EXCEL_FILE = "portfolio.xlsx"
SHEET_NAME = "포트폴리오"

COLUMNS_BASE = ["종목명", "자산유형", "플랫폼", "통화", "매입금액(원)", "평가금액(원)",
                "평가손익(원)", "수익률(%)", "비중(%)", "메모"]
COLUMNS_EXTENDED = COLUMNS_BASE + ["기초자산", "혼합비율"]


def load_portfolio(filepath: str = EXCEL_FILE) -> pd.DataFrame:
    if not Path(filepath).exists():
        raise FileNotFoundError(f"{filepath} 파일이 없습니다. create_template.py 먼저 실행하세요.")

    df = pd.read_excel(filepath, sheet_name=SHEET_NAME, header=0)
    COLUMNS = COLUMNS_EXTENDED if len(df.columns) >= 12 else COLUMNS_BASE
    df.columns = COLUMNS[:len(df.columns)]
    df = df.dropna(subset=["종목명"])
    df = df[df["종목명"].astype(str).str.strip() != ""].reset_index(drop=True)

    for col in ["매입금액(원)", "평가금액(원)"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # 자동 계산
    df["평가손익(원)"] = df["평가금액(원)"] - df["매입금액(원)"]
    df["수익률(%)"] = df.apply(
        lambda r: round(r["평가손익(원)"] / r["매입금액(원)"] * 100, 2) if r["매입금액(원)"] > 0 else 0,
        axis=1
    )
    total = df["평가금액(원)"].sum()
    df["비중(%)"] = (df["평가금액(원)"] / total * 100).round(2) if total > 0 else 0

    return df


def write_to_excel(df: pd.DataFrame, filepath: str = EXCEL_FILE):
    """계산된 값을 엑셀에 반영"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_NAME]
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=i, column=7, value=round(row["평가손익(원)"]))
        ws.cell(row=i, column=8, value=round(row["수익률(%)"], 2))
        ws.cell(row=i, column=9, value=round(row["비중(%)"], 2))
    wb.save(filepath)


def summarize(df: pd.DataFrame) -> dict:
    total_buy  = df["매입금액(원)"].sum()
    total_eval = df["평가금액(원)"].sum()
    total_profit = total_eval - total_buy
    total_return = round(total_profit / total_buy * 100, 2) if total_buy > 0 else 0

    by_type = df.groupby("자산유형").agg(
        평가금액합계=("평가금액(원)", "sum"),
        종목수=("종목명", "count"),
    ).copy()
    by_type["비중(%)"] = (by_type["평가금액합계"] / total_eval * 100).round(2)
    by_type = by_type.sort_values("비중(%)", ascending=False)

    return {
        "총_매입금액":    total_buy,
        "총_평가금액":    total_eval,
        "총_평가손익":    total_profit,
        "총_수익률(%)":   total_return,
        "자산유형별_분포": by_type,
        "종목수":         len(df),
    }


def print_report(df: pd.DataFrame):
    s = summarize(df)

    print("\n" + "=" * 62)
    print("📊 포트폴리오 현황")
    print("=" * 62)
    print(f"  종목 수     : {s['종목수']}개")
    print(f"  총 매입금액 : {s['총_매입금액']:>16,.0f} 원")
    print(f"  총 평가금액 : {s['총_평가금액']:>16,.0f} 원")
    print(f"  총 평가손익 : {s['총_평가손익']:>+16,.0f} 원")
    print(f"  총 수익률   : {s['총_수익률(%)']:>+.2f}%")

    print("\n" + "-" * 62)
    print("📂 자산유형별 분포")
    print("-" * 62)
    for asset_type, row in s["자산유형별_분포"].iterrows():
        bar = "█" * int(row["비중(%)"] / 2)
        print(f"  {asset_type:<8} | {int(row['종목수']):>2}종목 | "
              f"{row['평가금액합계']:>14,.0f}원 | {row['비중(%)']:>5.1f}% {bar}")

    print("\n" + "-" * 62)
    print("📋 종목별 상세 (비중 내림차순)")
    print("-" * 62)
    for _, row in df.sort_values("비중(%)", ascending=False).iterrows():
        sign = "▲" if row["수익률(%)"] >= 0 else "▼"
        print(f"  {row['종목명']:<15} [{row['자산유형']:<6}] "
              f"평가: {row['평가금액(원)']:>12,.0f}원 | "
              f"{sign}{abs(row['수익률(%)']):>5.2f}% | "
              f"비중: {row['비중(%)']:>5.2f}%")
    print("=" * 62)


def _calc_base_distribution(df: pd.DataFrame) -> dict[str, float]:
    """기초자산 컬럼 기반 실제 자산분포 계산 (혼합형 ETF 비율 분할 적용)"""
    import re
    base_data: dict[str, float] = {}
    for _, row in df.iterrows():
        base    = str(row.get("기초자산", "") or "").strip()
        mix_raw = row.get("혼합비율", "")
        mix_str = "" if (mix_raw is None or (isinstance(mix_raw, float) and __import__('math').isnan(mix_raw))) else str(mix_raw).strip()
        eval_   = row["평가금액(원)"]
        if not base:
            continue
        if base == "혼합" and mix_str:
            for part in mix_str.split("+"):
                m = re.match(r'(.+?)(\d+(?:\.\d+)?)%', part.strip())
                if m:
                    sub_base = m.group(1)
                    ratio    = float(m.group(2))
                    base_data[sub_base] = base_data.get(sub_base, 0) + eval_ * ratio / 100
        else:
            base_data[base] = base_data.get(base, 0) + eval_
    return base_data


def get_ai_prompt(df: pd.DataFrame) -> str:
    """AI 분석용 프롬프트 생성"""
    s = summarize(df)
    lines = [
        "아래는 현재 내 투자 포트폴리오 현황입니다.",
        "리밸런싱 시기와 방법, 위험 분산 측면에서 구체적인 조언을 해주세요.\n",
        "## 전체 요약",
        f"- 종목 수: {s['종목수']}개",
        f"- 총 평가금액: {s['총_평가금액']:,.0f}원",
        f"- 총 수익률: {s['총_수익률(%)']:+.2f}%\n",
        "## 자산유형별 비중 (상장 기준)",
    ]
    for asset_type, row in s["자산유형별_분포"].iterrows():
        lines.append(f"- {asset_type}: {row['비중(%)']:.1f}% ({int(row['종목수'])}종목, {row['평가금액합계']:,.0f}원)")

    # 기초자산 분포 (ETF 기초자산 기준, 혼합형 분할 적용)
    if "기초자산" in df.columns:
        total      = df["평가금액(원)"].sum()
        base_data  = _calc_base_distribution(df)
        if base_data:
            lines.append("\n## 기초자산별 실제 분포 (혼합형 ETF 비율 분할 적용)")
            for base, amount in sorted(base_data.items(), key=lambda x: -x[1]):
                pct = amount / total * 100
                lines.append(f"- {base}: {pct:.1f}% ({amount:,.0f}원)")

    lines.append("\n## 종목별 현황")
    for _, row in df.sort_values("비중(%)", ascending=False).iterrows():
        base = str(row.get("기초자산", "") or "").strip()
        mix  = row.get("혼합비율", "")
        mix  = "" if (mix is None or (isinstance(mix, float) and __import__('math').isnan(mix))) else str(mix).strip()
        base_info = f" | 기초자산: {base}" + (f" ({mix})" if mix else "") if base else ""
        lines.append(
            f"- {row['종목명']} ({row['자산유형']}, {row['통화']}): "
            f"평가금액 {row['평가금액(원)']:,.0f}원 | "
            f"수익률 {row['수익률(%)']:+.2f}% | "
            f"비중 {row['비중(%)']:.2f}%{base_info}"
        )
    return "\n".join(lines)


if __name__ == "__main__":
    df = load_portfolio()
    print_report(df)
    write_to_excel(df)
    print("\n✅ 엑셀 업데이트 완료")

    print("\n" + "=" * 62)
    print("🤖 AI 분석용 프롬프트")
    print("=" * 62)
    print(get_ai_prompt(df))
