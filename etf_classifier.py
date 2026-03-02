"""
ETF/종목 기초자산 분류기
- 포트폴리오 전체 종목을 Gemini가 분석하여 실제 노출 자산을 분류
- 컬럼 11: 기초자산 (미국주식, 국내채권 등 지역+유형 통합)
- 컬럼 12: 혼합비율 (혼합형 ETF만 — 예: 국내주식60%+미국채권40%)
- 혼합형 ETF는 실제 구성 비율로 평가금액을 분할하여 자산분포 계산

사용법: python3 etf_classifier.py
"""
import os
import json
import re
from dotenv import load_dotenv
from ai_client import get_client, get_fast_model
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from portfolio_reader import load_portfolio

load_dotenv()

EXCEL_FILE = "portfolio.xlsx"
SHEET_NAME = "포트폴리오"

COL_BASE = 11   # 기초자산
COL_MIX  = 12   # 혼합비율

# 허용되는 기초자산 값
VALID_BASE_ASSETS = {
    "미국주식", "미국채권",
    "국내주식", "국내채권",
    "인도주식", "일본주식", "중국주식",
    "글로벌주식", "신흥국주식",
    "원자재", "통화",
    "가상자산", "현금성", "혼합",
}

# 기초자산별 색상
BASE_COLORS = {
    "미국주식":   "D5E8D4",
    "미국채권":   "DAE8FC",
    "국내주식":   "DDEEFF",
    "국내채권":   "E8DAFC",
    "인도주식":   "FFF2CC",
    "일본주식":   "E1D5E7",
    "중국주식":   "FCE4D6",
    "글로벌주식": "C9E6F5",
    "신흥국주식": "F8CECC",
    "원자재":     "FFE6CC",
    "통화":       "EAD1DC",
    "가상자산":   "F9DBC6",
    "현금성":     "F2F2F2",
    "혼합":       "D9D9D9",
}

CLASSIFY_PROMPT = """
다음은 투자 포트폴리오의 종목 목록이야.
각 종목이 실제로 어느 자산에 투자하는지 분류해서 JSON 배열로만 응답해줘.

## 종목 목록
{items}

## 기초자산 분류 기준
단일 자산이면 아래 중 하나로 분류:
- 미국주식, 미국채권
- 국내주식, 국내채권
- 인도주식, 일본주식, 중국주식
- 글로벌주식, 신흥국주식
- 원자재, 통화, 가상자산, 현금성

혼합형 ETF (주식+채권, 여러 국가 혼합 등)이면:
- 기초자산: "혼합"
- 혼합비율: 실제 구성 자산과 비율 배열 (비율 합계 = 100)

## 분류 예시
- 삼성전자 (국내주식) → {{"name":"삼성전자", "기초자산":"국내주식", "혼합비율":[]}}
- TIGER 미국나스닥100 (국내ETF) → {{"name":"TIGER 미국나스닥100", "기초자산":"미국주식", "혼합비율":[]}}
- KODEX 200미국채혼합 (국내ETF) → {{"name":"KODEX 200미국채혼합", "기초자산":"혼합", "혼합비율":[{{"자산":"국내주식","비율":60}},{{"자산":"미국채권","비율":40}}]}}
- ACE 미국달러단기채권액티브 (국내ETF) → {{"name":"ACE 미국달러단기채권액티브", "기초자산":"미국채권", "혼합비율":[]}}
- TIGER 일본엔선물 (국내ETF) → {{"name":"TIGER 일본엔선물", "기초자산":"통화", "혼합비율":[]}}
- VOO (해외ETF) → {{"name":"VOO", "기초자산":"미국주식", "혼합비율":[]}}
- 비트코인 (가상자산) → {{"name":"비트코인", "기초자산":"가상자산", "혼합비율":[]}}
- TIGER 글로벌AI&로보틱스 → {{"name":"TIGER 글로벌AI&로보틱스", "기초자산":"글로벌주식", "혼합비율":[]}}
- KODEX 골드선물 → {{"name":"KODEX 골드선물", "기초자산":"원자재", "혼합비율":[]}}

## 주의사항
- KODEX/TIGER 등 국내 브랜드 ETF도 기초자산은 실제 투자 대상 기준 (국내상장≠국내주식)
- 혼합비율의 자산명은 위 기초자산 분류 값과 동일하게 사용
- 오직 JSON 배열만 출력 (마크다운, 코드블록, 설명 없이)

## 출력 형식
[
  {{"name": "종목명", "기초자산": "미국주식", "혼합비율": []}},
  {{"name": "혼합ETF명", "기초자산": "혼합", "혼합비율": [{{"자산": "국내주식", "비율": 60}}, {{"자산": "미국채권", "비율": 40}}]}}
]
"""


def classify_holdings(df, client) -> dict[str, dict]:
    """Gemini로 전체 종목 분류 → {종목명: {기초자산, 혼합비율}}"""
    items_text = "\n".join(
        f"- {row['종목명']} ({row['자산유형']})"
        for _, row in df.iterrows()
    )
    prompt = CLASSIFY_PROMPT.format(items=items_text)
    raw = client.generate(prompt, model=get_fast_model())
    raw = raw.strip().replace("```json", "").replace("```", "").strip()
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON 파싱 실패: {e}\n원본:\n{raw[:300]}")
        return {}

    result = {}
    for item in data:
        name  = str(item.get("name", "")).strip()
        base  = item.get("기초자산", "국내주식")
        mix   = item.get("혼합비율", [])
        if base not in VALID_BASE_ASSETS:
            base = "국내주식"
        # 혼합비율 검증: 비율 합이 100에 가까운지 확인
        if base == "혼합" and mix:
            total = sum(m.get("비율", 0) for m in mix)
            if total <= 0:
                base = "국내주식"
                mix  = []
        result[name] = {"기초자산": base, "혼합비율": mix}

    return result


def mix_ratio_to_str(mix: list) -> str:
    """혼합비율 배열 → 문자열 (예: 국내주식60%+미국채권40%)"""
    if not mix:
        return ""
    return "+".join(f"{m['자산']}{m['비율']}%" for m in mix)


def parse_mix_str(mix_str: str) -> list[dict]:
    """혼합비율 문자열 → 배열 파싱 (예: 국내주식60%+미국채권40% → [...])"""
    if not mix_str:
        return []
    parts = mix_str.split("+")
    result = []
    for part in parts:
        m = re.match(r'(.+?)(\d+(?:\.\d+)?)%', part.strip())
        if m:
            result.append({"자산": m.group(1), "비율": float(m.group(2))})
    return result


def ensure_header_columns(ws):
    """11, 12번 컬럼 헤더 설정"""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    for col, name, width in [(COL_BASE, "기초자산", 14), (COL_MIX, "혼합비율", 30)]:
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width


def write_classifications(ws, df, classifications: dict):
    """분류 결과를 포트폴리오 시트에 기록"""
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name:
            continue
        name = str(name).strip()
        cls  = classifications.get(name, {})
        base = cls.get("기초자산", "국내주식")
        mix  = cls.get("혼합비율", [])

        for col, val in [(COL_BASE, base), (COL_MIX, mix_ratio_to_str(mix))]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = center


def _expand_assets(df, classifications: dict) -> list[tuple]:
    """
    혼합형 ETF는 혼합비율에 따라 평가금액을 분할.
    반환: [(종목명, 자산유형, 기초자산, 평가금액), ...]
    """
    rows = []
    for _, row in df.iterrows():
        name  = str(row["종목명"]).strip()
        atype = row["자산유형"]
        eval_ = row["평가금액(원)"]
        cls   = classifications.get(name, {})
        base  = cls.get("기초자산", "국내주식")
        mix   = cls.get("혼합비율", [])

        if base == "혼합" and mix:
            for m in mix:
                rows.append((name, atype, m["자산"], eval_ * m["비율"] / 100))
        else:
            rows.append((name, atype, base, eval_))
    return rows


def make_distribution_sheet(wb, df, classifications: dict):
    """실제 자산분포 시트 생성 — 혼합형 ETF는 비율대로 분할"""
    sheet_name = "실제자산분포"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    total_eval = df["평가금액(원)"].sum()
    expanded   = _expand_assets(df, classifications)

    # 기초자산별 집계
    base_data: dict[str, float] = {}
    for _, _, base, amt in expanded:
        base_data[base] = base_data.get(base, 0) + amt

    # ── 제목 ──
    ws.merge_cells("A1:F1")
    ws["A1"].value     = "📊 실제 자산 분포 (기초자산 기준)"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # ── 기초자산별 분포 ──
    ws.cell(row=3, column=1, value="■ 기초자산별 분포").font = Font(bold=True, size=11)
    for col, h in enumerate(["기초자산", "평가금액(원)", "비중(%)", "시각화"], start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor="595959")
        cell.font      = Font(color="FFFFFF", bold=True)
        cell.alignment = center
        cell.border    = border

    r = 5
    for base, amount in sorted(base_data.items(), key=lambda x: -x[1]):
        pct  = amount / total_eval * 100
        bar  = "█" * int(pct / 2)
        fill = PatternFill("solid", fgColor=BASE_COLORS.get(base, "FFFFFF"))
        for col, val in enumerate([base, round(amount), round(pct, 2), bar], start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col != 4 else left
        r += 1

    # 합계 행
    for col, val in enumerate(["합계", round(total_eval), 100.0, ""], start=1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill      = PatternFill("solid", fgColor="D9D9D9")
        cell.font      = Font(bold=True)
        cell.border    = border
        cell.alignment = center
    r += 2

    # ── 종목별 상세 (혼합형은 분할 표시) ──
    ws.cell(row=r, column=1, value="■ 종목별 상세").font = Font(bold=True, size=11)
    r += 1
    detail_headers = ["종목명", "자산유형(표시)", "기초자산", "혼합비율", "평가금액(원)", "비중(%)"]
    for col, h in enumerate(detail_headers, start=1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor="595959")
        cell.font      = Font(color="FFFFFF", bold=True)
        cell.alignment = center
        cell.border    = border
    r += 1

    for _, row in df.sort_values("평가금액(원)", ascending=False).iterrows():
        name  = str(row["종목명"]).strip()
        atype = row["자산유형"]
        eval_ = row["평가금액(원)"]
        cls   = classifications.get(name, {})
        base  = cls.get("기초자산", "국내주식")
        mix   = cls.get("혼합비율", [])

        if base == "혼합" and mix:
            # 첫 줄: 종목명 + 혼합 표시
            mix_str = mix_ratio_to_str(mix)
            fill = PatternFill("solid", fgColor=BASE_COLORS.get("혼합", "D9D9D9"))
            for col, val in enumerate([name, atype, "혼합", mix_str, round(eval_), round(eval_ / total_eval * 100, 2)], start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center
            r += 1
            # 세부 분할 행
            for m in mix:
                sub_eval = eval_ * m["비율"] / 100
                sub_base = m["자산"]
                fill = PatternFill("solid", fgColor=BASE_COLORS.get(sub_base, "FFFFFF"))
                indent = f"  └ {sub_base} ({m['비율']}%)"
                for col, val in enumerate([indent, "", sub_base, "", round(sub_eval), round(sub_eval / total_eval * 100, 2)], start=1):
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = left if col == 1 else center
                r += 1
        else:
            fill = PatternFill("solid", fgColor=BASE_COLORS.get(base, "FFFFFF"))
            for col, val in enumerate([name, atype, base, "", round(eval_), round(eval_ / total_eval * 100, 2)], start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center
            r += 1

    for col, width in enumerate([22, 14, 14, 30, 16, 10], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return sheet_name


def print_summary(df, classifications: dict):
    total_eval = df["평가금액(원)"].sum()
    base_data: dict[str, float] = {}
    for _, _, base, amt in _expand_assets(df, classifications):
        base_data[base] = base_data.get(base, 0) + amt

    print("\n" + "=" * 55)
    print("📊 실제 자산 분포 (기초자산 기준, 혼합형 분할 적용)")
    print("=" * 55)
    for base, amount in sorted(base_data.items(), key=lambda x: -x[1]):
        pct = amount / total_eval * 100
        bar = "█" * int(pct / 2)
        print(f"  {base:<10} | {amount:>14,.0f}원 | {pct:>5.1f}% {bar}")
    print("=" * 55)


def main():
    print("\n📂 포트폴리오 로드 중...")
    df = load_portfolio()
    print(f"  → {len(df)}개 종목")

    client = get_client()

    print("\n🤖 Gemini로 기초자산 분류 중...")
    classifications = classify_holdings(df, client)

    if not classifications:
        print("❌ 분류 실패")
        return

    print(f"  → {len(classifications)}개 종목 분류 완료")

    print("\n📋 분류 결과:")
    for name, cls in classifications.items():
        base = cls["기초자산"]
        mix  = cls["혼합비율"]
        if mix:
            print(f"  {name:<25} → {base} ({mix_ratio_to_str(mix)})")
        else:
            print(f"  {name:<25} → {base}")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ensure_header_columns(ws)
    write_classifications(ws, df, classifications)
    make_distribution_sheet(wb, df, classifications)
    wb.save(EXCEL_FILE)

    print_summary(df, classifications)
    print(f"\n✅ portfolio.xlsx 업데이트 완료")
    print(f"   - 포트폴리오 시트: 기초자산 / 혼합비율 컬럼 갱신")
    print(f"   - '실제자산분포' 시트 재생성")


if __name__ == "__main__":
    main()
