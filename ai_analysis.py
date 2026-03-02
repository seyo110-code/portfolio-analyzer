"""
포트폴리오 AI 분석 및 리밸런싱 제안
- 포트폴리오 데이터를 Gemini에 전달
- Gemini가 질문 → 사용자 답변 → 리밸런싱 제안
- 제안 결과를 portfolio.xlsx의 새 시트에 기록

사용법: python3 ai_analysis.py
"""
import os
import json
from dotenv import load_dotenv
from ai_client import get_client, get_pro_model
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from portfolio_reader import load_portfolio, summarize, get_ai_prompt, print_report, write_to_excel, _calc_base_distribution
from etf_classifier import BASE_COLORS, parse_mix_str

load_dotenv()

EXCEL_FILE = "portfolio.xlsx"


# ─── Gemini 설정 ─────────────────────────────────────────────



# ─── 프롬프트 ────────────────────────────────────────────────

ANALYSIS_PROMPT = """{portfolio}

---
위 포트폴리오를 분석해줘. 아래 순서로 해줘:

1. **현황 분석**: 자산배분 현황, 집중 리스크, 눈에 띄는 특이사항 3~5줄로 요약
2. **질문**: 리밸런싱 제안을 위해 꼭 필요한 질문을 3~5개만 해줘

질문 형식 예시:
Q1. 투자 목표가 무엇인가요? (예: 자산 증식, 노후 대비, 단기 수익 등)
Q2. 리스크 허용 수준은 어느 정도인가요? (공격적 / 중립적 / 보수적)
Q3. ...

분석과 질문만 해줘. 아직 리밸런싱 제안은 하지 마.
"""

REBALANCE_PROMPT = """{portfolio}

---
사용자 답변:
{answers}

---
위 포트폴리오와 사용자 답변을 바탕으로 리밸런싱을 제안해줘.

반드시 아래 JSON 형식으로만 응답해줘 (마크다운, 설명 없이 순수 JSON만):

{{
  "summary": "리밸런싱 제안 요약 (3~5줄)",
  "items": [
    {{
      "name": "종목명",
      "current_weight": 현재비중(숫자),
      "proposed_weight": 제안비중(숫자),
      "action": "매수 또는 매도 또는 유지",
      "comment": "이 종목에 대한 한줄 코멘트"
    }}
  ],
  "new_items": [
    {{
      "name": "신규 추천 종목명",
      "asset_type": "자산유형",
      "proposed_weight": 제안비중(숫자),
      "comment": "추천 이유"
    }}
  ]
}}

주의사항:
- items는 현재 보유 중인 종목 전체 포함
- proposed_weight 합계는 반드시 100이 되도록
- new_items는 새로 편입을 추천하는 종목 (없으면 빈 배열)
- 비중은 소수점 1자리까지
"""


# ─── 대화 ────────────────────────────────────────────────────

def run_analysis(client, portfolio_prompt: str) -> str:
    print("\n" + "=" * 62)
    print(f"🤖 분석 중... ({get_pro_model()})")
    print("=" * 62)
    return client.generate(ANALYSIS_PROMPT.format(portfolio=portfolio_prompt), model=get_pro_model())


ANSWERS_FILE = Path("portfolio_answers.json")


def load_saved_answers() -> dict | None:
    if ANSWERS_FILE.exists():
        try:
            return json.loads(ANSWERS_FILE.read_text())
        except Exception:
            pass
    return None


def save_answers(answers: str):
    from datetime import datetime
    data = {
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "answers": answers,
    }
    ANSWERS_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2))


def _input_answers() -> str:
    print("   각 질문에 답 입력 후 Enter. 모든 답변이 끝나면 빈 줄에서 Enter 두 번.")
    print("-" * 62 + "\n")
    lines = []
    empty_count = 0
    while True:
        line = input()
        if line == "":
            empty_count += 1
            if empty_count >= 2:
                break
        else:
            empty_count = 0
            lines.append(line)
    return "\n".join(lines)


def collect_answers(analysis_text: str) -> str:
    print("\n" + analysis_text)
    print("\n" + "=" * 62)
    print("💬 답변 입력")
    print("=" * 62)

    saved = load_saved_answers()
    if saved:
        print(f"\n  📁 이전 답변 기록 ({saved['saved_at']})")
        print("  " + "-" * 58)
        for line in saved["answers"].splitlines():
            print(f"  {line}")
        print("  " + "-" * 58)
        print("  1. 이전 답변 그대로 사용")
        print("  2. 이전 답변 수정 후 사용")
        print("  3. 새로 입력")
        choice = input("\n  선택 (Enter = 그대로 사용): ").strip()

        if choice == "" or choice == "1":
            print("  → 이전 답변 사용\n")
            return saved["answers"]

        if choice == "2":
            print("\n  이전 답변을 수정하세요. 수정 완료 후 빈 줄에서 Enter 두 번.")
            print(f"\n--- 이전 답변 ---\n{saved['answers']}\n-----------------\n")
            answers = _input_answers()
            if not answers.strip():
                answers = saved["answers"]
            save_answers(answers)
            return answers

    # 새로 입력 (choice == "3" 또는 저장된 답변 없음)
    answers = _input_answers()
    if answers.strip():
        save_answers(answers)
    return answers


FEEDBACK_PROMPT = """{portfolio}

---
사용자 답변:
{answers}

---
이전 제안 요약:
{prev_summary}

---
사용자 피드백:
{feedback}

---
위 피드백을 반영하여 리밸런싱 제안을 수정해줘.
반드시 아래 JSON 형식으로만 응답해줘 (마크다운, 설명 없이 순수 JSON만):

{{
  "summary": "수정된 리밸런싱 제안 요약 (3~5줄)",
  "items": [
    {{
      "name": "종목명",
      "current_weight": 현재비중(숫자),
      "proposed_weight": 제안비중(숫자),
      "action": "매수 또는 매도 또는 유지",
      "comment": "이 종목에 대한 한줄 코멘트"
    }}
  ],
  "new_items": [
    {{
      "name": "신규 추천 종목명",
      "asset_type": "자산유형",
      "proposed_weight": 제안비중(숫자),
      "comment": "추천 이유"
    }}
  ]
}}

주의사항:
- items는 현재 보유 중인 종목 전체 포함
- proposed_weight 합계는 반드시 100이 되도록
- new_items는 새로 편입을 추천하는 종목 (없으면 빈 배열)
- 비중은 소수점 1자리까지
"""


def _parse_rebalance(raw: str) -> dict:
    raw = raw.strip().replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON 파싱 실패: {e}")
        print(f"원본 응답:\n{raw[:300]}")
        return {}


def run_rebalance(client, portfolio_prompt: str, answers: str) -> dict:
    print(f"\n🤖 리밸런싱 제안 생성 중... ({get_pro_model()})")
    prompt = REBALANCE_PROMPT.format(portfolio=portfolio_prompt, answers=answers)
    return _parse_rebalance(client.generate(prompt, model=get_pro_model()))


def run_rebalance_with_feedback(client, portfolio_prompt: str, answers: str,
                                 prev_result: dict, feedback: str) -> dict:
    print(f"\n🤖 피드백 반영하여 재생성 중... ({get_pro_model()})")
    prompt = FEEDBACK_PROMPT.format(
        portfolio=portfolio_prompt,
        answers=answers,
        prev_summary=prev_result.get("summary", ""),
        feedback=feedback,
    )
    return _parse_rebalance(client.generate(prompt, model=get_pro_model()))


# ─── 엑셀 시트 작성 ──────────────────────────────────────────

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_proposal_sheet(df, result: dict, filepath: str = EXCEL_FILE):
    wb = openpyxl.load_workbook(filepath)
    sheet_name = f"리밸런싱제안_{datetime.now().strftime('%m%d')}"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # ── 스타일 상수 ──
    blue_fill    = PatternFill("solid", fgColor="2F5496")
    dark_fill    = PatternFill("solid", fgColor="595959")
    summary_fill = PatternFill("solid", fgColor="EBF3FB")
    buy_fill     = PatternFill("solid", fgColor="D5E8D4")
    sell_fill    = PatternFill("solid", fgColor="FCE4D6")
    hold_fill    = PatternFill("solid", fgColor="F2F2F2")
    new_fill     = PatternFill("solid", fgColor="FFF2CC")
    total_fill   = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    border = make_border()

    total_eval = df["평가금액(원)"].sum()
    r = 1  # 현재 행 포인터

    # ══════════════════════════════════════════════
    # 1. 제목 + AI 요약
    # ══════════════════════════════════════════════
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(row=r, column=1, value=f"📊 리밸런싱 제안  ({datetime.now().strftime('%Y-%m-%d')})")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = blue_fill;  c.alignment = center
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(f"A{r}:I{r+2}")
    c = ws.cell(row=r, column=1, value=result.get("summary", ""))
    c.fill = summary_fill;  c.alignment = left
    for i in range(3):
        ws.row_dimensions[r + i].height = 18
    r += 4  # 요약 3행 + 여백 1행

    # ══════════════════════════════════════════════
    # 2. 기초자산 분포 비교 (현재 vs 제안)
    # ══════════════════════════════════════════════
    if "기초자산" in df.columns:
        cur_base  = _calc_base_distribution(df)
        prop_base: dict[str, float] = {}

        prop_map = {i["name"]: i.get("proposed_weight", 0)
                    for i in list(result.get("items", [])) +
                              [{**n, "action": "신규매수"} for n in result.get("new_items", [])]}

        for _, row in df.iterrows():
            base    = str(row.get("기초자산", "") or "").strip()
            mix_str = row.get("혼합비율", "")
            mix_str = "" if (mix_str is None or (isinstance(mix_str, float) and __import__('math').isnan(mix_str))) else str(mix_str).strip()
            prop_w  = prop_map.get(row["종목명"], row["비중(%)"])
            prop_amt = total_eval * prop_w / 100
            if base == "혼합" and mix_str:
                for m in parse_mix_str(mix_str):
                    prop_base[m["자산"]] = prop_base.get(m["자산"], 0) + prop_amt * m["비율"] / 100
            elif base:
                prop_base[base] = prop_base.get(base, 0) + prop_amt

        for new_item in result.get("new_items", []):
            label    = new_item.get("asset_type") or new_item.get("name", "기타")
            prop_amt = total_eval * new_item.get("proposed_weight", 0) / 100
            prop_base[label] = prop_base.get(label, 0) + prop_amt

        # 섹션 제목
        ws.merge_cells(f"A{r}:I{r}")
        c = ws.cell(row=r, column=1, value="■ 기초자산 분포 비교 (현재 vs 제안)")
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = blue_fill;  c.alignment = center
        ws.row_dimensions[r].height = 22
        r += 1

        # 헤더
        dist_headers = ["기초자산", "현재금액(원)", "현재(%)", "→",
                        "제안금액(원)", "제안(%)", "변화(%p)",
                        "현재 시각화", "제안 시각화"]
        for col, h in enumerate(dist_headers, start=1):
            c = ws.cell(row=r, column=col, value=h)
            c.fill = dark_fill;  c.font = Font(color="FFFFFF", bold=True)
            c.alignment = center;  c.border = border
        ws.row_dimensions[r].height = 20
        r += 1

        all_bases = sorted(set(list(cur_base) + list(prop_base)),
                           key=lambda x: -(cur_base.get(x, 0) + prop_base.get(x, 0)))
        for base in all_bases:
            cur_amt  = cur_base.get(base, 0)
            prop_amt = prop_base.get(base, 0)
            cur_pct  = cur_amt  / total_eval * 100
            prop_pct = prop_amt / total_eval * 100
            delta    = round(prop_pct - cur_pct, 2)
            fill     = PatternFill("solid", fgColor=BASE_COLORS.get(base, "FFFFFF"))
            cur_bar  = "█" * max(1, int(cur_pct  / 2))
            prop_bar = "█" * max(1, int(prop_pct / 2))
            for col, val in enumerate([
                base, round(cur_amt), round(cur_pct, 2), "→",
                round(prop_amt), round(prop_pct, 2), delta, cur_bar, prop_bar
            ], start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = fill;  c.border = border
                c.alignment = left if col in (8, 9) else center
            ws.row_dimensions[r].height = 18
            r += 1
        r += 1  # 여백

    # ══════════════════════════════════════════════
    # 3. 종목별 비중 조절
    # ══════════════════════════════════════════════
    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(row=r, column=1, value="■ 종목별 비중 조절")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = blue_fill;  c.alignment = center
    ws.row_dimensions[r].height = 22
    r += 1

    ticker_headers = ["종목명", "기초자산", "현재비중(%)", "제안비중(%)",
                      "현재금액(원)", "제안금액(원)", "변경금액(원)", "조치", "AI 코멘트"]
    for col, h in enumerate(ticker_headers, start=1):
        c = ws.cell(row=r, column=col, value=h)
        c.fill = dark_fill;  c.font = Font(color="FFFFFF", bold=True)
        c.alignment = center;  c.border = border
    ws.row_dimensions[r].height = 20
    r += 1

    item_map = {i["name"]: i for i in result.get("items", [])}

    for _, row in df.sort_values("비중(%)", ascending=False).iterrows():
        name       = row["종목명"]
        proposal   = item_map.get(name, {})
        prop_w     = proposal.get("proposed_weight", row["비중(%)"])
        action     = proposal.get("action", "유지")
        comment    = proposal.get("comment", "")
        cur_amt    = round(row["평가금액(원)"])
        prop_amt   = round(total_eval * prop_w / 100)
        change     = prop_amt - cur_amt
        fill       = buy_fill if action == "매수" else sell_fill if action == "매도" else hold_fill

        base = str(row.get("기초자산", "") or "").strip() or row["자산유형"]
        for col, val in enumerate([name, base, row["비중(%)"], prop_w,
                                    cur_amt, prop_amt, change, action, comment], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill;  c.border = border
            c.alignment = left if col == 9 else center
        ws.row_dimensions[r].height = 18
        r += 1

    # 신규 추천
    new_items = result.get("new_items", [])
    if new_items:
        ws.merge_cells(f"A{r}:I{r}")
        c = ws.cell(row=r, column=1, value="▼ 신규 편입 추천")
        c.fill = dark_fill;  c.font = Font(color="FFFFFF", bold=True);  c.alignment = center
        ws.row_dimensions[r].height = 18
        r += 1
        for item in new_items:
            prop_w   = item.get("proposed_weight", 0)
            prop_amt = round(total_eval * prop_w / 100)
            for col, val in enumerate([item.get("name"), item.get("asset_type"), "-", prop_w,
                                        "-", prop_amt, prop_amt, "신규매수", item.get("comment", "")], start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = new_fill;  c.border = border
                c.alignment = left if col == 9 else center
            ws.row_dimensions[r].height = 18
            r += 1

    # 합계
    for col, val in enumerate(["합계", "", round(df["비중(%)"].sum(), 1), "",
                                round(total_eval), "", "", "", ""], start=1):
        c = ws.cell(row=r, column=col, value=val if val != "" else None)
        c.fill = total_fill;  c.border = border;  c.alignment = center
        if val != "":
            c.font = Font(bold=True)
    ws.row_dimensions[r].height = 20

    # ── 컬럼 너비 (두 섹션 공용 최적값) ──
    col_widths = [22, 14, 11, 11, 16, 16, 14, 20, 36]
    for col, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # 첫 행 고정
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return sheet_name


# ─── 메인 ────────────────────────────────────────────────────

def main():
    print("\n📂 포트폴리오 데이터 로드 중...")
    df = load_portfolio()
    s  = summarize(df)
    portfolio_prompt = get_ai_prompt(df)

    print_report(df)
    write_to_excel(df)

    client = get_client()

    # 1단계: 분석 + 질문
    analysis = run_analysis(client, portfolio_prompt)

    # 2단계: 사용자 답변 수집
    answers = collect_answers(analysis)
    if not answers.strip():
        print("⚠️ 답변이 없어 기본 분석으로 진행합니다.")
        answers = "특별한 선호 없음. 균형 잡힌 포트폴리오를 원함."

    # 3단계: 리밸런싱 제안 + 피드백 루프
    result = run_rebalance(client, portfolio_prompt, answers)
    if not result:
        print("❌ 리밸런싱 제안 생성 실패")
        return

    round_num = 1
    while True:
        print("\n" + "=" * 62)
        print(f"💡 리밸런싱 제안 요약 (Round {round_num})")
        print("=" * 62)
        print(result.get("summary", ""))
        print("\n" + "-" * 62)
        print("종목별 제안:")
        for item in result.get("items", []):
            action = item.get("action", "유지")
            icon   = "📈" if action == "매수" else "📉" if action == "매도" else "➡️"
            print(f"  {icon} {item['name']:<25} {item.get('current_weight', 0):>5.1f}% → {item.get('proposed_weight', 0):>5.1f}%  {item.get('comment','')}")
        if result.get("new_items"):
            print("신규 추천:")
            for item in result.get("new_items", []):
                print(f"  ➕ {item['name']:<25} → {item.get('proposed_weight', 0):>5.1f}%  {item.get('comment','')}")

        print("\n" + "=" * 62)
        feedback = input("💬 피드백 입력 (수정 원하면 입력 / 확정은 Enter): ").strip()
        if not feedback:
            break

        result = run_rebalance_with_feedback(client, portfolio_prompt, answers, result, feedback)
        if not result:
            print("❌ 재생성 실패, 이전 제안을 사용합니다.")
            break
        round_num += 1

    # 최종 엑셀 저장
    sheet_name = write_proposal_sheet(df, result)
    print(f"\n✅ '{sheet_name}' 시트가 portfolio.xlsx에 추가되었습니다.")
    print("   매수(연두) / 매도(연빨) / 유지(회색) / 신규(노랑) 색상으로 표시됩니다.")


if __name__ == "__main__":
    main()
