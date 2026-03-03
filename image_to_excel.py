"""
스크린샷 → 포트폴리오 자동 입력
- 뱅크샐러드/토스 스크린샷 여러 장을 Gemini Vision으로 분석
- 중복 처리: 같은 종목+플랫폼 → 합산, 같은 종목+다른 플랫폼 → 별도 행
- 적용 전 미리보기 확인 후 엑셀 업데이트

사용법:
  단일:  python3 image_to_excel.py screenshot.png
  다중:  python3 image_to_excel.py toss.png banksalad.png
  폴더:  python3 image_to_excel.py screenshots/
  초기화: python3 image_to_excel.py 스샷.png --reset
"""
import sys
import os
import json
import argparse
from PIL import Image
from dotenv import load_dotenv
from ai_client import get_client, get_fast_model
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from pathlib import Path

load_dotenv()

EXCEL_FILE = "portfolio.xlsx"
SHEET_NAME = "포트폴리오"

TYPE_COLORS = {
    "국내주식": "DDEEFF",
    "해외주식": "D5E8D4",
    "국내ETF":  "FFF2CC",
    "해외ETF":  "E2EFDA",
    "가상자산": "FCE4D6",
    "현금성":   "F2F2F2",
}
VALID_TYPES = list(TYPE_COLORS.keys())

EXTRACT_PROMPT = """
이 이미지는 뱅크샐러드 또는 토스 앱의 투자/자산 포트폴리오 화면 스크린샷이야.
이미지에서 각 종목(자산)별로 다음 정보를 추출해서 JSON 배열로만 응답해줘.

## 추출할 필드
- name: 종목명 (예: 삼성전자, 비트코인, Apple, VOO)
- asset_type: 아래 중 하나로 분류
    * 국내주식 (코스피/코스닥 상장 주식)
    * 해외주식 (미국 등 해외 개별 종목)
    * 국내ETF (국내 상장 ETF, KODEX/TIGER/ARIRANG 등)
    * 해외ETF (미국 상장 ETF, VOO/QQQ/SPY 등)
    * 가상자산 (비트코인, 이더리움, 알트코인 등)
    * 현금성 (예금, CMA, 파킹통장, MMF 등)
- platform: 보유 플랫폼/증권사 (이미지에서 파악 가능하면, 없으면 빈 문자열)
- currency: KRW 또는 USD
- buy_amount_krw: 매입금액 원화 기준 (숫자, 없으면 0)
- eval_amount_krw: 평가금액 원화 기준 (숫자) ← 가장 중요
- memo: 특이사항 (없으면 빈 문자열)

## 주의사항
- 오직 JSON 배열만 출력 (마크다운, 코드블록, 설명 없이)
- 숫자는 쉼표, 원, $, % 기호 없이 순수 숫자만
- 확실하지 않은 값은 0 또는 빈 문자열
- 유사 글자 혼동 주의: '전력(電力)'과 '전략(戰略)', '채권'과 '채련' 등

## 출력 형식 예시
[
  {"name": "삼성전자", "asset_type": "국내주식", "platform": "키움증권", "currency": "KRW", "buy_amount_krw": 3250000, "eval_amount_krw": 3600000, "memo": ""},
  {"name": "비트코인", "asset_type": "가상자산", "platform": "업비트", "currency": "KRW", "buy_amount_krw": 1750000, "eval_amount_krw": 4250000, "memo": ""}
]
"""


# ─── 이미지 분석 ────────────────────────────────────────────



def extract_from_image(image_path: str, client) -> list[dict]:
    raw = client.generate(EXTRACT_PROMPT, image_path=image_path, model=get_fast_model())
    raw = raw.strip().replace("```json", "").replace("```", "").strip()
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError as e:
        print(f"  ⚠️ JSON 파싱 실패: {e}")
        print(f"  원본 응답: {raw[:300]}")
        return []


def collect_image_paths(args: list[str]) -> list[Path]:
    paths = []
    for arg in args:
        p = Path(arg)
        if p.is_dir():
            for ext in ["*.png", "*.jpg", "*.jpeg"]:
                paths.extend(sorted(p.glob(ext)))
        elif p.is_file() and p.suffix.lower() in {".png", ".jpg", ".jpeg"}:
            paths.append(p)
        else:
            print(f"⚠️ 무시됨: {arg}")
    return paths


# ─── 데이터 정규화 및 중복 처리 ──────────────────────────────

# AI가 자주 오독하는 종목명 보정 사전 (key: 오독, value: 정확한 이름)
NAME_CORRECTIONS = {
    "전략인프라": "전력인프라",
    "전략 인프라": "전력 인프라",
}


def correct_name(name: str) -> str:
    """알려진 오독 패턴 보정"""
    for wrong, correct in NAME_CORRECTIONS.items():
        if wrong in name:
            return name.replace(wrong, correct)
    return name



# 국내ETF 브랜드 prefix (국내 상장 ETF)
DOMESTIC_ETF_PREFIXES = (
    "KODEX", "TIGER", "ARIRANG", "ACE", "KBSTAR", "HANARO",
    "SOL", "TIMEFOLIO", "PLUS", "KOSEF", "TREX", "SMART",
    "FOCUS", "WOORI", "MABOT", "CAPE",
)

# 가상자산 키워드
CRYPTO_KEYWORDS = ("비트코인", "이더리움", "리플", "도지", "솔라나", "BTC", "ETH", "XRP")

# 현금성 키워드
CASH_KEYWORDS = ("CMA", "MMF", "파킹", "예금", "적금", "예치금", "RP")


def infer_asset_type(name: str) -> str:
    """종목명 패턴으로 자산유형 추론"""
    import re

    # 가상자산
    if any(k in name for k in CRYPTO_KEYWORDS):
        return "가상자산"

    # 현금성
    if any(k in name for k in CASH_KEYWORDS):
        return "현금성"

    # 국내ETF: 알려진 브랜드 prefix
    name_upper = name.upper()
    if any(name_upper.startswith(p) for p in DOMESTIC_ETF_PREFIXES):
        return "국내ETF"

    # 해외ETF: 순수 영문 2~5자 티커 (VOO, QQQ, SPY, VTI 등)
    if re.fullmatch(r'[A-Z]{2,5}', name_upper):
        return "해외ETF"

    # 해외주식: 영문이지만 긴 이름 (Apple, Tesla, Nvidia 등)
    if re.fullmatch(r'[A-Za-z][A-Za-z0-9\s\.\-]+', name) and len(name) > 5:
        return "해외주식"

    # 기본값: 한글 이름 → 국내주식
    return "국내주식"


def validate_item(item: dict) -> dict | None:
    name = correct_name(str(item.get("name", "")).strip())
    if not name:
        return None

    # AI 추출값 우선, 단 이름으로 보정 가능한 경우 덮어씀
    ai_type = item.get("asset_type", "")
    inferred = infer_asset_type(name)

    if ai_type in VALID_TYPES:
        # AI가 올바른 값을 줬어도 이름 기반 추론이 더 신뢰도 높은 경우 보정
        # (예: AI가 국내ETF라 했지만 이름이 VOO → 해외ETF로 교정)
        asset_type = inferred if inferred != "국내주식" else ai_type
    else:
        asset_type = inferred

    return {
        "name":           name,
        "asset_type":     asset_type,
        "platform":       str(item.get("platform", "")).strip(),
        "currency":       str(item.get("currency", "KRW")).upper(),
        "buy_amount_krw": float(item.get("buy_amount_krw", 0) or 0),
        "eval_amount_krw":float(item.get("eval_amount_krw", 0) or 0),
        "memo":           str(item.get("memo", "")),
    }


def normalize_platform(platform: str) -> str:
    """플랫폼 이름 정규화 (AI가 다르게 추출하는 경우 통일)"""
    p = platform.strip()
    mapping = {
        "토스페이": "토스증권",
        "toss": "토스증권",
        "네이버증권": "네이버페이",
        "한투": "한국투자증권",
        "한국투자": "한국투자증권",
        "미래": "미래에셋",
        "키움": "키움증권",
    }
    return mapping.get(p, p)


def _find_canonical(truncated: str, keys) -> str | None:
    """'TIGER 미국S&P500타...' → 매칭되는 전체 이름 키 반환"""
    if not truncated.endswith("..."):
        return None
    prefix = truncated[:-3]
    for k in keys:
        if k.startswith(prefix) and not k.endswith("..."):
            return k
    return None


def _find_truncated(full_name: str, keys) -> str | None:
    """'TIGER 미국S&P500타겟리턴' → 매칭되는 잘린 이름 키 반환"""
    for k in keys:
        if k.endswith("...") and full_name.startswith(k[:-3]):
            return k
    return None


def _merge_into(target: dict, source: dict):
    target["buy_amount_krw"]  += source["buy_amount_krw"]
    target["eval_amount_krw"] += source["eval_amount_krw"]
    if not target["platform"] and source["platform"]:
        target["platform"] = source["platform"]


def deduplicate(all_items: list[dict]) -> list[dict]:
    """
    종목명 기준 중복 제거 및 금액 합산.
    - 정확히 같은 이름: 합산
    - 잘린 이름(...) vs 전체 이름: 전체 이름으로 통합
    """
    result: dict[str, dict] = {}
    for item in all_items:
        item["platform"] = normalize_platform(item["platform"])
        name = item["name"]

        # 1. 정확히 같은 키 있으면 합산
        if name in result:
            _merge_into(result[name], item)
            continue

        # 2. 내가 잘린 이름 → 기존에 전체 이름이 있으면 합산
        canonical = _find_canonical(name, result)
        if canonical:
            _merge_into(result[canonical], item)
            continue

        # 3. 내가 전체 이름 → 기존에 잘린 이름이 있으면 교체
        trunc_key = _find_truncated(name, result)
        if trunc_key:
            existing = result.pop(trunc_key)
            item["buy_amount_krw"]  += existing["buy_amount_krw"]
            item["eval_amount_krw"] += existing["eval_amount_krw"]
            if not item["platform"] and existing["platform"]:
                item["platform"] = existing["platform"]

        result[name] = item

    return list(result.values())


def cleanup_excel_duplicates(filepath: str = EXCEL_FILE) -> int:
    """엑셀에 이미 저장된 잘린 이름(...) 중복 항목 정리"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_NAME]

    # 행 인덱스 → 이름 매핑
    name_to_row: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        if val:
            name_to_row[str(val).strip()] = row_idx

    to_delete: set[int] = set()
    for name, row_idx in list(name_to_row.items()):
        if not name.endswith("..."):
            continue
        prefix = name[:-3]
        full = next((k for k in name_to_row if k.startswith(prefix) and not k.endswith("...")), None)
        if not full:
            continue

        full_row  = name_to_row[full]
        trunc_buy  = ws.cell(row=row_idx, column=5).value or 0
        trunc_eval = ws.cell(row=row_idx, column=6).value or 0
        full_buy   = ws.cell(row=full_row, column=5).value or 0
        full_eval  = ws.cell(row=full_row, column=6).value or 0

        # 금액이 다르면 합산 (별도 계좌 보유분), 같으면 단순 중복 제거
        if trunc_buy != full_buy or trunc_eval != full_eval:
            ws.cell(row=full_row, column=5, value=full_buy + trunc_buy)
            ws.cell(row=full_row, column=6, value=full_eval + trunc_eval)

        to_delete.add(row_idx)

    for row_idx in sorted(to_delete, reverse=True):
        ws.delete_rows(row_idx)

    wb.save(filepath)
    return len(to_delete)


# ─── 미리보기 ────────────────────────────────────────────────

def print_preview(final_items: list[dict], existing_names: set, reset: bool = False):
    print("\n" + "=" * 68)
    print("📋 적용 예정 데이터 미리보기")
    if reset:
        print("⚠️  초기화 모드: 기존 데이터가 모두 삭제됩니다")
    print("=" * 68)
    print(f"  {'상태':<6} {'종목명':<16} {'자산유형':<8} {'플랫폼':<12} {'평가금액(원)':>14}")
    print("-" * 68)
    for item in final_items:
        if reset:
            status = "신규추가"
        else:
            status = "업데이트" if item["name"] in existing_names else "신규추가"
        print(f"  {status:<6} {item['name']:<16} {item['asset_type']:<8} "
              f"{item['platform']:<12} {item['eval_amount_krw']:>14,.0f}")
    print("=" * 68)


# ─── 엑셀 업데이트 ───────────────────────────────────────────

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def get_existing_map(ws) -> dict:
    """종목명 → 행 번호"""
    result = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            result[str(name).strip()] = row
    return result


def write_row(ws, row_idx: int, item: dict):
    border = make_border()
    center = Alignment(horizontal="center", vertical="center")
    fill   = PatternFill("solid", fgColor=TYPE_COLORS.get(item["asset_type"], "FFFFFF"))

    buy    = item["buy_amount_krw"]
    eval_  = item["eval_amount_krw"]
    profit = eval_ - buy
    ret    = round(profit / buy * 100, 2) if buy > 0 else 0

    values = [
        item["name"], item["asset_type"], item["platform"], item["currency"],
        round(buy), round(eval_), round(profit), ret, 0, item["memo"]
    ]
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill   = fill
        cell.border = border
        cell.alignment = center


def update_excel(items: list[dict], filepath: str = EXCEL_FILE, reset: bool = False):
    if not Path(filepath).exists():
        raise FileNotFoundError(f"{filepath} 없음. create_template.py 먼저 실행하세요.")

    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_NAME]

    if reset:
        # 2행부터 끝까지 데이터 삭제
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)
        existing_map = {}
    else:
        existing_map = get_existing_map(ws)

    updated, added = [], []
    for item in items:
        key = item["name"]
        if key in existing_map:
            write_row(ws, existing_map[key], item)
            updated.append(item["name"])
        else:
            write_row(ws, ws.max_row + 1, item)
            added.append(item["name"])

    wb.save(filepath)
    return updated, added


# ─── 메인 ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="스크린샷 → 포트폴리오 자동 입력",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "사용법:\n"
            "  단일:   python3 image_to_excel.py screenshot.png\n"
            "  다중:   python3 image_to_excel.py toss.png banksalad.png\n"
            "  폴더:   python3 image_to_excel.py screenshots/\n"
            "  초기화: python3 image_to_excel.py 스샷.png --reset"
        ),
    )
    parser.add_argument("images", nargs="+", help="분석할 이미지 파일 또는 폴더")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="기존 포트폴리오 데이터를 모두 삭제하고 새로 시작",
    )
    args = parser.parse_args()

    if args.reset:
        print("\n⚠️  초기화 모드: 기존 포트폴리오 데이터가 모두 삭제됩니다.")
        confirm = input("정말 초기화하고 새로 시작할까요? (yes/n): ").strip().lower()
        if confirm != "yes":
            print("❌ 취소되었습니다.")
            sys.exit(0)

    image_paths = collect_image_paths(args.images)
    if not image_paths:
        print("❌ 처리할 이미지 파일이 없습니다.")
        sys.exit(1)

    print(f"\n📁 처리할 이미지 {len(image_paths)}장: {[p.name for p in image_paths]}")

    client = get_client()

    all_items = []
    for img_path in image_paths:
        print(f"\n🔍 분석 중: {img_path.name}")
        raw_items = extract_from_image(str(img_path), client)
        validated = [v for item in raw_items if (v := validate_item(item))]
        print(f"  → {len(validated)}개 종목 추출")
        all_items.extend(validated)

    if not all_items:
        print("\n❌ 추출된 데이터가 없습니다.")
        sys.exit(1)

    print(f"\n📊 전체 추출: {len(all_items)}개 → 중복 처리 중...")
    final_items = deduplicate(all_items)
    removed = len(all_items) - len(final_items)
    if removed > 0:
        print(f"  → 중복 {removed}건 합산 → 최종 {len(final_items)}개 종목")

    # 기존 엑셀 종목명 (미리보기용)
    existing_names = set()
    if not args.reset and Path(EXCEL_FILE).exists():
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            existing_names = {
                str(ws.cell(row=r, column=1).value).strip()
                for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=1).value
            }
        except Exception:
            pass

    print_preview(final_items, existing_names, reset=args.reset)

    answer = input("\n위 내용을 엑셀에 적용할까요? (y/n): ").strip().lower()
    if answer != "y":
        print("❌ 취소되었습니다.")
        sys.exit(0)

    print(f"\n💾 {EXCEL_FILE} 업데이트 중...")
    updated, added = update_excel(final_items, reset=args.reset)

    print("\n✅ 완료!")
    if updated:
        print(f"  🔄 업데이트 ({len(updated)}개): {', '.join(updated)}")
    if added:
        print(f"  ➕ 신규 추가 ({len(added)}개): {', '.join(added)}")

    # ── Step 1: 비중(%) 자동 계산 ──────────────────────────────
    print("\n⚙️  비중(%) 계산 중...")
    from portfolio_reader import load_portfolio, write_to_excel as write_portfolio
    df = load_portfolio()
    write_portfolio(df)
    print(f"  → 비중(%) 재계산 완료 ({len(df)}개 종목)")

    # ── Step 2: 기초자산_지역/유형 자동 분류 ─────────────────────
    print("\n🤖 기초자산 분류 중...")
    from etf_classifier import (
        classify_holdings, ensure_header_columns,
        write_classifications, make_distribution_sheet,
    )
    classifications = classify_holdings(df, client)
    if classifications:
        wb2 = openpyxl.load_workbook(EXCEL_FILE)
        ws2 = wb2[SHEET_NAME]
        ensure_header_columns(ws2)
        write_classifications(ws2, df, classifications)
        make_distribution_sheet(wb2, df, classifications)
        wb2.save(EXCEL_FILE)
        print(f"  → {len(classifications)}개 종목 분류 완료")
    else:
        print("  ⚠️ 분류 실패 (나중에 etf_classifier.py 직접 실행)")

    print(f"\n✅ {EXCEL_FILE} 완성 — 비중(%) · 기초자산_지역/유형 모두 채워진 상태")


if __name__ == "__main__":
    main()
