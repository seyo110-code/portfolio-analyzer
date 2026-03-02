"""
포트폴리오 엑셀 템플릿 생성 스크립트
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_portfolio_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "포트폴리오"

    headers = [
        ("종목명",       20),
        ("자산유형",     14),
        ("플랫폼",       14),
        ("통화",          8),
        ("매입금액(원)", 16),
        ("평가금액(원)", 16),
        ("평가손익(원)", 16),
        ("수익률(%)",    12),
        ("비중(%)",      10),
        ("메모",         20),
        ("기초자산",     14),
        ("혼합비율",     30),
    ]

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 25

    # 데이터 없이 헤더만 생성 - image_to_excel.py로 스크린샷에서 채워넣기

    # 입력 안내 시트
    ws2 = wb.create_sheet("입력안내")
    guide = [
        ("항목",         "설명",                                                    "예시"),
        ("종목명",       "종목 또는 자산 이름",                                      "삼성전자, 비트코인, VOO"),
        ("자산유형",     "국내주식/해외주식/국내ETF/해외ETF/가상자산/현금성 중 하나", "국내주식"),
        ("플랫폼",       "보유 증권사/거래소",                                        "키움증권, 업비트"),
        ("통화",         "KRW 또는 USD",                                             "USD"),
        ("매입금액(원)", "총 매입금액 원화 기준",                                     "3250000"),
        ("평가금액(원)", "현재 평가금액 원화 기준 ← 가장 중요",                       "3600000"),
        ("평가손익(원)", "자동계산",                                                  ""),
        ("수익률(%)",    "자동계산",                                                  ""),
        ("비중(%)",      "자동계산",                                                  ""),
        ("메모",         "자유 메모",                                                 "헤지용"),
        ("",             "",                                                          ""),
        ("주의",         "USD 자산의 금액은 원화로 환산해서 입력하세요.",              ""),
    ]
    for r, row in enumerate(guide, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)
            ws2.column_dimensions[get_column_letter(c)].width = [15, 55, 30][c - 1]

    wb.save("portfolio.xlsx")
    print("✅ portfolio.xlsx 생성 완료!")
    print("📋 헤더만 생성됐습니다. screenshots/ 폴더에 스크린샷을 넣고 image_to_excel.py를 실행하세요.")


if __name__ == "__main__":
    create_portfolio_template()
